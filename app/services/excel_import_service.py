"""
Excel file import service for processing .xlsx and .xls files.
Handles complex Excel spreadsheets with multiple sheets and formatting.

Created by Team Sigma (Data Processing & Import) - Sprint 8
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import logging
from typing import Dict, List, Optional, Tuple
from datetime import datetime

logger = logging.getLogger('import.excel')


class ExcelImportService:
    """Specialized Excel file import service."""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls']
        self.date_formats = [
            '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d',
            '%m-%d-%Y', '%d-%m-%Y', '%m/%d/%y', '%d/%m/%y',
            '%b %d, %Y', '%B %d, %Y', '%d %b %Y', '%d %B %Y'
        ]
    
    def parse_excel_file(self, file_path, sheet_name=None):
        """
        Parse Excel file and return DataFrame.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet to parse (None for first sheet)
            
        Returns:
            pandas.DataFrame: Parsed Excel data
        """
        
        try:
            # Load workbook to get sheet information
            workbook_info = self.get_workbook_info(file_path)
            
            # Determine which sheet to use
            if sheet_name is None:
                sheet_name = workbook_info['sheets'][0]['name']
            
            logger.info(f"Parsing Excel sheet: {sheet_name}", extra={
                'file_path': file_path,
                'sheet_name': sheet_name,
                'total_sheets': len(workbook_info['sheets'])
            })
            
            # Read Excel file
            if file_path.endswith('.xlsx'):
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            else:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
            
            # Clean and process DataFrame
            df = self._clean_excel_dataframe(df)
            
            # Validate structure
            self._validate_excel_structure(df, sheet_name)
            
            logger.info(f"Excel file parsed successfully", extra={
                'rows': len(df),
                'columns': len(df.columns),
                'sheet_name': sheet_name
            })
            
            return df
            
        except Exception as e:
            logger.error(f"Excel parsing failed: {str(e)}", extra={
                'file_path': file_path,
                'sheet_name': sheet_name
            })
            raise ValueError(f"Failed to parse Excel file: {str(e)}")
    
    def get_workbook_info(self, file_path):
        """
        Get information about Excel workbook structure.
        
        Returns:
            dict: Workbook information including sheets and data ranges
        """
        
        try:
            if file_path.endswith('.xlsx'):
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                sheets = []
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Calculate data range
                    max_row = sheet.max_row
                    max_col = sheet.max_column
                    
                    # Estimate data density
                    sample_cells = 0
                    filled_cells = 0
                    
                    for row in range(1, min(11, max_row + 1)):  # Sample first 10 rows
                        for col in range(1, min(11, max_col + 1)):  # Sample first 10 columns
                            cell = sheet.cell(row=row, column=col)
                            sample_cells += 1
                            if cell.value is not None:
                                filled_cells += 1
                    
                    data_density = (filled_cells / sample_cells) if sample_cells > 0 else 0
                    
                    sheets.append({
                        'name': sheet_name,
                        'max_row': max_row,
                        'max_col': max_col,
                        'data_density': round(data_density, 2),
                        'estimated_transactions': max_row - 1 if max_row > 1 else 0
                    })
                
                workbook.close()
                
            else:  # .xls file
                # For .xls files, use pandas to get basic info
                excel_file = pd.ExcelFile(file_path)
                sheets = []
                
                for sheet_name in excel_file.sheet_names:
                    df_sample = pd.read_excel(file_path, sheet_name=sheet_name, nrows=10)
                    
                    sheets.append({
                        'name': sheet_name,
                        'max_row': 'unknown',
                        'max_col': len(df_sample.columns),
                        'data_density': 'unknown',
                        'estimated_transactions': 'unknown'
                    })
            
            return {
                'file_type': 'Excel',
                'sheets': sheets,
                'recommended_sheet': self._recommend_best_sheet(sheets)
            }
            
        except Exception as e:
            logger.error(f"Failed to get workbook info: {str(e)}")
            raise ValueError(f"Unable to read Excel file structure: {str(e)}")
    
    def _recommend_best_sheet(self, sheets):
        """Recommend the best sheet for transaction data."""
        
        if len(sheets) == 1:
            return sheets[0]['name']
        
        # Score sheets based on likelihood of containing transaction data
        best_sheet = None
        best_score = 0
        
        for sheet in sheets:
            score = 0
            
            # Prefer sheets with transaction-like names
            transaction_keywords = ['transaction', 'trans', 'statement', 'activity', 'history']
            sheet_name_lower = sheet['name'].lower()
            
            for keyword in transaction_keywords:
                if keyword in sheet_name_lower:
                    score += 10
            
            # Prefer sheets with reasonable data density
            if isinstance(sheet['data_density'], float):
                if 0.3 <= sheet['data_density'] <= 0.8:  # Good data density
                    score += 5
            
            # Prefer sheets with reasonable number of rows
            if isinstance(sheet['estimated_transactions'], int):
                if 10 <= sheet['estimated_transactions'] <= 10000:
                    score += 3
            
            if score > best_score:
                best_score = score
                best_sheet = sheet['name']
        
        return best_sheet or sheets[0]['name']
    
    def _clean_excel_dataframe(self, df):
        """Clean and prepare Excel DataFrame for processing."""
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Reset index after dropping rows
        df = df.reset_index(drop=True)
        
        # Clean column names
        df.columns = [str(col).strip() for col in df.columns]
        
        # Remove unnamed columns (often from Excel formatting)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        
        # Handle merged cells (fill forward)
        for column in df.columns:
            df[column] = df[column].fillna(method='ffill')
        
        return df
    
    def _validate_excel_structure(self, df, sheet_name):
        """Validate Excel DataFrame structure."""
        
        if df.empty:
            raise ValueError(f"Excel sheet '{sheet_name}' is empty")
        
        if len(df.columns) < 3:
            raise ValueError(f"Excel sheet '{sheet_name}' must have at least 3 columns")
        
        # Check for reasonable amount of data
        if len(df) > 100000:  # 100k rows
            raise ValueError(f"Excel sheet '{sheet_name}' has too many rows ({len(df)}). Maximum 100,000 rows.")
        
        logger.info(f"Excel structure validated", extra={
            'sheet_name': sheet_name,
            'rows': len(df),
            'columns': len(df.columns)
        })
    
    def detect_excel_column_types(self, df):
        """Detect column types in Excel data."""
        
        column_types = {}
        
        for column in df.columns:
            # Get non-null sample values
            sample_values = df[column].dropna().head(20)
            
            if len(sample_values) == 0:
                column_types[column] = 'empty'
                continue
            
            # Convert to string for analysis
            sample_strings = sample_values.astype(str)
            
            # Detect column type
            if self._is_excel_date_column(sample_values):
                column_types[column] = 'date'
            elif self._is_excel_amount_column(sample_strings):
                column_types[column] = 'amount'
            elif self._is_excel_description_column(sample_strings):
                column_types[column] = 'description'
            elif self._is_excel_account_column(sample_strings):
                column_types[column] = 'account'
            else:
                column_types[column] = 'text'
        
        return column_types
    
    def _is_excel_date_column(self, sample_values):
        """Check if Excel column contains dates."""
        
        date_count = 0
        
        for value in sample_values:
            # Check if value is already a datetime object (Excel date)
            if isinstance(value, (pd.Timestamp, datetime)):
                date_count += 1
                continue
            
            # Try parsing as string date
            value_str = str(value).strip()
            for date_format in self.date_formats:
                try:
                    datetime.strptime(value_str, date_format)
                    date_count += 1
                    break
                except ValueError:
                    continue
        
        return date_count >= len(sample_values) * 0.7  # 70% must be dates
    
    def _is_excel_amount_column(self, sample_strings):
        """Check if Excel column contains monetary amounts."""
        
        amount_count = 0
        
        for value in sample_strings:
            # Clean value for amount detection
            clean_value = value.replace('$', '').replace(',', '').replace('€', '')
            clean_value = clean_value.replace('£', '').replace('¥', '').strip()
            
            # Handle parentheses for negative amounts
            if clean_value.startswith('(') and clean_value.endswith(')'):
                clean_value = clean_value[1:-1]
            
            try:
                float(clean_value)
                amount_count += 1
            except ValueError:
                continue
        
        return amount_count >= len(sample_strings) * 0.8  # 80% must be numbers
    
    def _is_excel_description_column(self, sample_strings):
        """Check if Excel column contains transaction descriptions."""
        
        # Description characteristics
        avg_length = sum(len(str(v)) for v in sample_strings) / len(sample_strings)
        unique_ratio = len(set(sample_strings)) / len(sample_strings)
        
        # Check for description-like content
        description_indicators = 0
        description_keywords = [
            'payment', 'purchase', 'transfer', 'deposit', 'withdrawal',
            'fee', 'charge', 'refund', 'credit', 'debit', 'atm'
        ]
        
        for value in sample_strings:
            value_lower = str(value).lower()
            if any(keyword in value_lower for keyword in description_keywords):
                description_indicators += 1
        
        # High uniqueness, reasonable length, and transaction keywords
        return (unique_ratio > 0.6 and avg_length > 8) or description_indicators >= len(sample_strings) * 0.3
    
    def _is_excel_account_column(self, sample_strings):
        """Check if Excel column contains account information."""
        
        # Account patterns
        account_patterns = [
            r'\d{4,}',  # Account numbers
            r'.*checking.*', r'.*saving.*', r'.*credit.*',
            r'.*account.*', r'.*acct.*'
        ]
        
        import re
        account_indicators = 0
        
        for value in sample_strings:
            value_lower = str(value).lower()
            for pattern in account_patterns:
                if re.search(pattern, value_lower):
                    account_indicators += 1
                    break
        
        return account_indicators >= len(sample_strings) * 0.5  # 50% must match patterns
    
    def suggest_column_mappings(self, df):
        """Suggest column mappings for Excel data."""
        
        column_types = self.detect_excel_column_types(df)
        mappings = {}
        
        # Find best candidates for each required field
        date_candidates = [col for col, type_ in column_types.items() if type_ == 'date']
        amount_candidates = [col for col, type_ in column_types.items() if type_ == 'amount']
        description_candidates = [col for col, type_ in column_types.items() if type_ == 'description']
        
        # Assign mappings
        if date_candidates:
            mappings[date_candidates[0]] = 'date'
        
        if amount_candidates:
            mappings[amount_candidates[0]] = 'amount'
        
        if description_candidates:
            mappings[description_candidates[0]] = 'description'
        
        # If we couldn't auto-detect, suggest based on column names
        for column in df.columns:
            col_lower = column.lower()
            
            if 'date' in col_lower and 'date' not in mappings.values():
                mappings[column] = 'date'
            elif 'amount' in col_lower and 'amount' not in mappings.values():
                mappings[column] = 'amount'
            elif any(word in col_lower for word in ['description', 'memo', 'detail']) and 'description' not in mappings.values():
                mappings[column] = 'description'
        
        return mappings
    
    def preview_excel_data(self, file_path, sheet_name=None, max_rows=10):
        """
        Preview Excel data for user review.
        
        Returns:
            dict: Preview data with structure information
        """
        
        try:
            # Get workbook info
            workbook_info = self.get_workbook_info(file_path)
            
            # Parse specified sheet or first sheet
            if sheet_name is None:
                sheet_name = workbook_info['recommended_sheet']
            
            # Read sample data
            if file_path.endswith('.xlsx'):
                df_sample = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows, engine='openpyxl')
            else:
                df_sample = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows, engine='xlrd')
            
            # Clean sample data
            df_sample = self._clean_excel_dataframe(df_sample)
            
            # Detect column types
            column_types = self.detect_excel_column_types(df_sample)
            
            # Suggest mappings
            suggested_mappings = self.suggest_column_mappings(df_sample)
            
            # Calculate data quality score
            quality_score = self._calculate_data_quality_score(df_sample)
            
            return {
                'workbook_info': workbook_info,
                'selected_sheet': sheet_name,
                'preview_data': {
                    'columns': list(df_sample.columns),
                    'sample_rows': df_sample.head(5).to_dict('records'),
                    'total_rows': len(df_sample),
                    'column_types': column_types,
                    'suggested_mappings': suggested_mappings
                },
                'data_quality': {
                    'score': quality_score,
                    'issues': self._identify_data_issues(df_sample)
                }
            }
            
        except Exception as e:
            logger.error(f"Excel preview failed: {str(e)}")
            raise ValueError(f"Unable to preview Excel file: {str(e)}")
    
    def _calculate_data_quality_score(self, df):
        """Calculate data quality score for Excel data."""
        
        score = 100
        
        # Penalize for missing data
        missing_percentage = (df.isnull().sum().sum() / (len(df) * len(df.columns))) * 100
        score -= missing_percentage * 0.5
        
        # Penalize for duplicate rows
        duplicate_percentage = (df.duplicated().sum() / len(df)) * 100
        score -= duplicate_percentage
        
        # Penalize for inconsistent data types
        for column in df.columns:
            unique_types = df[column].dropna().apply(type).nunique()
            if unique_types > 1:
                score -= 5  # Mixed data types in column
        
        return max(0, min(100, score))
    
    def _identify_data_issues(self, df):
        """Identify potential data quality issues."""
        
        issues = []
        
        # Check for missing data
        missing_data = df.isnull().sum()
        for column, missing_count in missing_data.items():
            if missing_count > 0:
                percentage = (missing_count / len(df)) * 100
                if percentage > 10:
                    issues.append({
                        'type': 'missing_data',
                        'column': column,
                        'description': f"Column '{column}' has {percentage:.1f}% missing values",
                        'severity': 'high' if percentage > 50 else 'medium'
                    })
        
        # Check for duplicate rows
        duplicate_count = df.duplicated().sum()
        if duplicate_count > 0:
            percentage = (duplicate_count / len(df)) * 100
            issues.append({
                'type': 'duplicate_rows',
                'description': f"{duplicate_count} duplicate rows found ({percentage:.1f}%)",
                'severity': 'medium' if percentage > 5 else 'low'
            })
        
        # Check for inconsistent data types
        for column in df.columns:
            sample_values = df[column].dropna().head(20)
            if len(sample_values) > 0:
                types = sample_values.apply(type).unique()
                if len(types) > 1:
                    issues.append({
                        'type': 'mixed_data_types',
                        'column': column,
                        'description': f"Column '{column}' has mixed data types",
                        'severity': 'low'
                    })
        
        # Check for very large or very small values that might be errors
        numeric_columns = df.select_dtypes(include=['number']).columns
        for column in numeric_columns:
            values = df[column].dropna()
            if len(values) > 0:
                mean_val = values.mean()
                std_val = values.std()
                
                # Check for outliers (values > 3 standard deviations from mean)
                outliers = values[abs(values - mean_val) > 3 * std_val]
                if len(outliers) > 0:
                    issues.append({
                        'type': 'outlier_values',
                        'column': column,
                        'description': f"Column '{column}' has {len(outliers)} potential outlier values",
                        'severity': 'low'
                    })
        
        return issues
    
    def convert_excel_to_transactions(self, df, column_mappings):
        """Convert Excel DataFrame to transaction records using mappings."""
        
        transactions = []
        
        for index, row in df.iterrows():
            try:
                transaction_data = {}
                
                # Map columns to transaction fields
                for excel_column, transaction_field in column_mappings.items():
                    if excel_column in df.columns and transaction_field != 'ignore':
                        value = row[excel_column]
                        
                        # Process value based on target field
                        if transaction_field == 'date':
                            transaction_data['date'] = self._process_excel_date(value)
                        elif transaction_field == 'amount':
                            transaction_data['amount'] = self._process_excel_amount(value)
                        elif transaction_field == 'description':
                            transaction_data['description'] = self._process_excel_text(value)
                        elif transaction_field == 'currency':
                            transaction_data['currency'] = self._process_excel_currency(value)
                        else:
                            transaction_data[transaction_field] = str(value) if value is not None else ''
                
                # Add row metadata
                transaction_data['row_number'] = index + 1
                transaction_data['raw_data'] = row.to_dict()
                
                transactions.append(transaction_data)
                
            except Exception as e:
                logger.warning(f"Failed to process Excel row {index + 1}: {str(e)}")
                # Continue processing other rows
                continue
        
        return transactions
    
    def _process_excel_date(self, value):
        """Process Excel date value."""
        
        if pd.isna(value):
            return None
        
        # If already a datetime object
        if isinstance(value, (pd.Timestamp, datetime)):
            return value.date()
        
        # Try parsing as string
        value_str = str(value).strip()
        for date_format in self.date_formats:
            try:
                return datetime.strptime(value_str, date_format).date()
            except ValueError:
                continue
        
        raise ValueError(f"Unable to parse date: {value}")
    
    def _process_excel_amount(self, value):
        """Process Excel amount value."""
        
        if pd.isna(value):
            return None
        
        # If already a number
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        
        # Clean string value
        value_str = str(value).strip()
        
        # Handle parentheses for negative amounts
        if value_str.startswith('(') and value_str.endswith(')'):
            value_str = '-' + value_str[1:-1]
        
        # Remove currency symbols and formatting
        clean_value = value_str.replace('$', '').replace(',', '')
        clean_value = clean_value.replace('€', '').replace('£', '').replace('¥', '')
        
        try:
            return Decimal(clean_value)
        except (ValueError, InvalidOperation):
            raise ValueError(f"Unable to parse amount: {value}")
    
    def _process_excel_text(self, value):
        """Process Excel text value."""
        
        if pd.isna(value):
            return ''
        
        return str(value).strip()
    
    def _process_excel_currency(self, value):
        """Process Excel currency value."""
        
        if pd.isna(value):
            return 'USD'
        
        currency = str(value).strip().upper()
        
        # Map common currency symbols to codes
        currency_mapping = {
            '$': 'USD',
            '€': 'EUR', 
            '£': 'GBP',
            '¥': 'JPY',
            'C$': 'CAD',
            'A$': 'AUD'
        }
        
        return currency_mapping.get(currency, currency if len(currency) == 3 else 'USD')